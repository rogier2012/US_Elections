from openpyxl import load_workbook
import json
wb2 = load_workbook('Votes2.xlsx')

ws = wb2.active

data = []

for i in range(5,57):
    state = dict()
    state['name'] = ws["A" + str(i)].value
    state['population'] = ws["B" + str(i)].value
    state['seats'] = ws["C" + str(i)].value
    state['TCJ'] = ws["E" + str(i)].value
    state['TJC'] = ws["F" + str(i)].value
    state['CTJ'] = ws["G" + str(i)].value
    state['CJT'] = ws["H" + str(i)].value
    state['JTC'] = ws["I" + str(i)].value
    state['JCT'] = ws["J" + str(i)].value
    data.append(state)

with open('votes.json', 'w') as outfile:
    json.dump(data, outfile, indent=4)


